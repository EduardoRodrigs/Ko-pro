import io
from fastapi import FastAPI, Depends, Request, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from datetime import datetime
import csv
import math
from openpyxl import load_workbook
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError

from database import engine, Base, init_db, get_db, Cliente, MetaMensal, ProdutoMeta, PositivacaoDinamica, HistoricoChat, HistoricoPositivacaoMensal

# Geolocator setup using Nominatim
geolocator = Nominatim(user_agent="andina_pro_geocoder_v2")

def geocode_address(address_str):
    try:
        location = geolocator.geocode(address_str, timeout=3)
        if location:
            return location.latitude, location.longitude
    except (GeocoderTimedOut, GeocoderServiceError):
        pass
    return None, None

def geocode_missing_clients(client_ids: list, db_session_maker):
    db = db_session_maker()
    try:
        geocoded_count = 0
        for cid in client_ids:
            if geocoded_count >= 3:
                break
            c = db.query(Cliente).filter(Cliente.cod_cliente == cid).first()
            if c and (c.latitude is None or c.longitude is None):
                full_address = f"{c.endereco or ''}, {c.bairro or ''}, {c.cidade or ''} - RJ"
                try:
                    lat, lng = geocode_address(full_address)
                    if lat and lng:
                        c.latitude = lat
                        c.longitude = lng
                        db.add(c)
                        geocoded_count += 1
                except Exception:
                    pass
        if geocoded_count > 0:
            db.commit()
    except Exception as e:
        print("Erro no geocoding em background:", e)
    finally:
        db.close()

def haversine_distance(lat1, lon1, lat2, lon2):
    r = 6371  # radius of Earth in km
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    
    a = math.sin(delta_phi / 2)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return r * c

def parse_float(val):
    try:
        if val is None:
            return None
        val_str = str(val).strip().replace(',', '.')
        return float(val_str) if val_str else None
    except ValueError:
        return None

from openai import OpenAI
import os
import re

openai_client = None
api_key = os.getenv("OPENAI_API_KEY")
if api_key:
    openai_client = OpenAI(api_key=api_key)

def format_markdown(text_str):
    # Simple regex for bold **text**
    text_str = re.sub(r'\*\*(.*?)\*\*', r'<strong>\1</strong>', text_str)
    # Simple regex for italics *text*
    text_str = re.sub(r'\*(.*?)\*', r'<em>\1</em>', text_str)
    # Convert linebreaks
    text_str = text_str.replace('\n', '<br>')
    return text_str

def render_chat_bubbles(user_text, ia_text):
    user_html = f"""
    <div class="flex justify-end mb-4">
        <div class="bg-brand text-white rounded-2xl rounded-tr-none px-4 py-2.5 max-w-[85%] text-xs shadow-sm leading-relaxed">
            {user_text}
        </div>
    </div>
    """
    formatted_ia = format_markdown(ia_text)
    ia_html = f"""
    <div class="flex justify-start mb-4">
        <div class="bg-white text-gray-800 rounded-2xl rounded-tl-none px-4 py-2.5 max-w-[85%] text-xs shadow-sm border border-gray-100 leading-relaxed">
            {formatted_ia}
        </div>
    </div>
    """
    return user_html + ia_html

def get_current_and_prev_months():
    now = datetime.now()
    curr_yyyymm = now.strftime("%Y-%m") # e.g. '2026-06'
    
    # Calculate previous month
    if now.month == 1:
        prev_month = 12
        prev_year = now.year - 1
    else:
        prev_month = now.month - 1
        prev_year = now.year
        
    prev_mmyyyy = f"{prev_month:02d}_{prev_year}" # e.g. '05_2026'
    return curr_yyyymm, prev_mmyyyy

app = FastAPI(title="Andina Pro")

# Initialize database

# Initialize database
init_db()

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# Templates
templates = Jinja2Templates(directory="templates")

def get_current_month():
    return datetime.now().strftime("%Y-%m")

@app.get("/", response_class=HTMLResponse)
async def read_dashboard(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/config", response_class=HTMLResponse)
async def read_config(request: Request):
    return templates.TemplateResponse("config.html", {"request": request})

@app.get("/metas", response_class=HTMLResponse)
async def read_metas(request: Request):
    return templates.TemplateResponse("metas.html", {"request": request})

@app.get("/chat", response_class=HTMLResponse)
async def read_chat(request: Request, rota: str = None, db: Session = Depends(get_db)):
    query = db.query(HistoricoChat)
    if rota:
        query = query.filter(HistoricoChat.rota_ativa == rota)
    historico = query.order_by(HistoricoChat.data_hora.asc()).all()
    return templates.TemplateResponse("chat.html", {"request": request, "historico": historico})

@app.get("/cliente/{cod_cliente}", response_class=HTMLResponse)
async def read_cliente(request: Request, cod_cliente: str):
    is_htmx = request.headers.get("HX-Request") == "true"
    template_name = "cliente_content.html" if is_htmx else "cliente.html"
    return templates.TemplateResponse(template_name, {
        "request": request,
        "cod_cliente": cod_cliente,
        "is_htmx": is_htmx
    })

# --- API ENDPOINTS ---

@app.post("/api/scan-routes")
async def scan_routes(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        filename = file.filename.lower()
        
        unique_routes = set()
        
        if filename.endswith('.xlsx'):
            wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            sheet = wb.active
            
            # Read first row
            headers_row = next(sheet.iter_rows(max_row=1, values_only=True), None)
            if not headers_row:
                return JSONResponse(status_code=400, content={"message": "Excel file is empty"})
                
            headers = [str(cell).strip() if cell is not None else '' for cell in headers_row]
            
            route_idx = None
            for idx, col in enumerate(headers):
                if col.strip().upper() == 'NOVA ROTA':
                    route_idx = idx + 1 # 1-based index for openpyxl
                    break
                    
            if route_idx is None:
                return JSONResponse(status_code=400, content={"message": "Não foi possível encontrar a coluna 'Nova Rota' na planilha Excel."})
                
            for row in sheet.iter_rows(min_row=2, min_col=route_idx, max_col=route_idx, values_only=True):
                if row:
                    val = str(row[0] or '').strip()
                    if val:
                        unique_routes.add(val)
        else:
            # Fallback to CSV
            content_str = contents.decode('utf-8-sig')
            dialect = csv.excel
            dialect.delimiter = ';'
            reader = csv.DictReader(io.StringIO(content_str), dialect=dialect)
            if reader.fieldnames:
                reader.fieldnames = [col.strip() for col in reader.fieldnames]
                
            route_col = None
            if reader.fieldnames:
                for col in reader.fieldnames:
                    if col.strip().upper() == 'NOVA ROTA':
                        route_col = col
                        break
                        
            if route_col is None:
                return JSONResponse(status_code=400, content={"message": "Não foi possível encontrar a coluna 'Nova Rota' na planilha CSV."})
                
            for row in reader:
                val = str(row.get(route_col) or '').strip()
                if val:
                    unique_routes.add(val)
                    
        return {"routes": sorted(list(unique_routes))}
    except Exception as e:
        return JSONResponse(status_code=400, content={"message": f"Erro ao analisar rotas: {str(e)}"})

@app.post("/api/upload")
async def upload_csv(file: UploadFile = File(...), rota: str = Form(...), db: Session = Depends(get_db)):
    try:
        contents = await file.read()
        filename = file.filename.lower()
        
        rows_to_process = []
        
        if filename.endswith('.xlsx'):
            wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            sheet = wb.active
            
            # Read first row
            headers_row = next(sheet.iter_rows(max_row=1, values_only=True), None)
            if not headers_row:
                raise HTTPException(status_code=400, detail="Excel file is empty")
            
            # Extract and clean headers
            headers = [str(cell).strip() if cell is not None else '' for cell in headers_row]
            print("COLUNAS ENCONTRADAS NO EXCEL:", headers)
            
            # Find the "Nova Rota" column index
            route_idx = None
            for idx, col in enumerate(headers):
                if col.strip().upper() == 'NOVA ROTA':
                    route_idx = idx
                    break
            if route_idx is None:
                raise HTTPException(status_code=400, detail="Coluna 'Nova Rota' não encontrada no Excel.")
                    # Map Excel columns dynamically
            header_map = {}
            for idx, col in enumerate(headers):
                col_upper = col.strip().upper()
                if col_upper.startswith('COD CLIEN') or col_upper.startswith('COD CLIENT') or ('COD' in col_upper and 'CLIEN' in col_upper):
                    header_map['cod_cliente'] = idx
                elif col_upper == 'RAZAO SOCIAL':
                    header_map['razao_social'] = idx
                elif col_upper == 'ENDERECO':
                    header_map['endereco'] = idx
                elif col_upper == 'BAIRRO':
                    header_map['bairro'] = idx
                elif col_upper == 'CIDADE':
                    header_map['cidade'] = idx
                elif col_upper == 'CANAL RESUMIDO':
                    header_map['canal_resumido'] = idx
                elif col_upper in ('CLASSIFICAC', 'CLASSIFICACAO'):
                    header_map['classificacao'] = idx
                elif col_upper == 'NOVO DIA':
                    header_map['novo_dia'] = idx
                elif col_upper == 'NOVA SEMANA':
                    header_map['nova_semana'] = idx
                elif col_upper in ('LATITUDE KNVV', 'LATITUDE'):
                    header_map['latitude'] = idx
                elif col_upper in ('LONGITUDE KNVV', 'LONGITUDE'):
                    header_map['longitude'] = idx
            
            required_db_fields = ['cod_cliente', 'razao_social', 'endereco', 'bairro', 'cidade', 'canal_resumido', 'classificacao', 'novo_dia', 'nova_semana']
            missing_fields = [f for f in required_db_fields if f not in header_map]
            
            if missing_fields:
                print("ERRO DE COLUNAS EXCEL. Faltam:", missing_fields)
                raise HTTPException(status_code=400, detail=f"Excel is missing required columns: {missing_fields}")
                
            # Build standard dict list
            for row_cells in sheet.iter_rows(min_row=2, values_only=True):
                # Skip if row is empty or doesn't have the elements we need
                if not row_cells or len(row_cells) <= max(header_map.values()) or len(row_cells) <= route_idx:
                    continue
                row_route = str(row_cells[route_idx] or '').strip()
                if row_route != rota:
                    continue
                rows_to_process.append({
                    'cod_cliente': str(row_cells[header_map['cod_cliente']] or '').strip(),
                    'razao_social': str(row_cells[header_map['razao_social']] or '').strip(),
                    'endereco': str(row_cells[header_map['endereco']] or '').strip(),
                    'bairro': str(row_cells[header_map['bairro']] or '').strip(),
                    'cidade': str(row_cells[header_map['cidade']] or '').strip(),
                    'canal_resumido': str(row_cells[header_map['canal_resumido']] or '').strip(),
                    'classificacao': str(row_cells[header_map['classificacao']] or '').strip(),
                    'novo_dia': str(row_cells[header_map['novo_dia']] or '').strip(),
                    'nova_semana': str(row_cells[header_map['nova_semana']] or '').strip(),
                    'latitude': parse_float(row_cells[header_map['latitude']]) if 'latitude' in header_map else None,
                    'longitude': parse_float(row_cells[header_map['longitude']]) if 'longitude' in header_map else None
                })
        else:
            # Fallback to CSV
            # Decode using utf-8-sig to automatically handle and remove BOM if present
            content_str = contents.decode('utf-8-sig')
            
            # Force semicolon and strip whitespaces from field names
            dialect = csv.excel
            dialect.delimiter = ';'
                
            reader = csv.DictReader(io.StringIO(content_str), dialect=dialect)
            
            # Strip column names to avoid issues with spaces
            if reader.fieldnames:
                reader.fieldnames = [col.strip() for col in reader.fieldnames]
                
            print("COLUNAS ENCONTRADAS NO CSV:", reader.fieldnames)
            
            route_col = None
            if reader.fieldnames:
                for col in reader.fieldnames:
                    if col.strip().upper() == 'NOVA ROTA':
                        route_col = col
                        break
            if route_col is None:
                raise HTTPException(status_code=400, detail="Coluna 'Nova Rota' não encontrada no CSV.")
            
            # Standardize and map headers
            header_map = {}
            if reader.fieldnames:
                for col in reader.fieldnames:
                    col_upper = col.strip().upper()
                    if col_upper.startswith('COD CLIEN') or col_upper.startswith('COD CLIENT') or ('COD' in col_upper and 'CLIEN' in col_upper):
                        header_map['cod_cliente'] = col
                    elif col_upper == 'RAZAO SOCIAL':
                        header_map['razao_social'] = col
                    elif col_upper == 'ENDERECO':
                        header_map['endereco'] = col
                    elif col_upper == 'BAIRRO':
                        header_map['bairro'] = col
                    elif col_upper == 'CIDADE':
                        header_map['cidade'] = col
                    elif col_upper == 'CANAL RESUMIDO':
                        header_map['canal_resumido'] = col
                    elif col_upper in ('CLASSIFICAC', 'CLASSIFICACAO'):
                        header_map['classificacao'] = col
                    elif col_upper == 'NOVO DIA':
                        header_map['novo_dia'] = col
                    elif col_upper == 'NOVA SEMANA':
                        header_map['nova_semana'] = col
                    elif col_upper in ('LATITUDE KNVV', 'LATITUDE'):
                        header_map['latitude'] = col
                    elif col_upper in ('LONGITUDE KNVV', 'LONGITUDE'):
                        header_map['longitude'] = col
 
            required_db_fields = ['cod_cliente', 'razao_social', 'endereco', 'bairro', 'cidade', 'canal_resumido', 'classificacao', 'novo_dia', 'nova_semana']
            missing_fields = [f for f in required_db_fields if f not in header_map]
            
            if missing_fields:
                print("ERRO DE COLUNAS CSV. Faltam:", missing_fields)
                raise HTTPException(status_code=400, detail=f"CSV is missing required columns: {missing_fields}")
                
            for row in reader:
                row_route = str(row.get(route_col) or '').strip()
                if row_route != rota:
                    continue
                rows_to_process.append({
                    'cod_cliente': str(row.get(header_map['cod_cliente']) or '').strip(),
                    'razao_social': str(row.get(header_map['razao_social']) or '').strip(),
                    'endereco': str(row.get(header_map['endereco']) or '').strip(),
                    'bairro': str(row.get(header_map['bairro']) or '').strip(),
                    'cidade': str(row.get(header_map['cidade']) or '').strip(),
                    'canal_resumido': str(row.get(header_map['canal_resumido']) or '').strip(),
                    'classificacao': str(row.get(header_map['classificacao']) or '').strip(),
                    'novo_dia': str(row.get(header_map['novo_dia']) or '').strip(),
                    'nova_semana': str(row.get(header_map['nova_semana']) or '').strip(),
                    'latitude': parse_float(row.get(header_map['latitude'])) if 'latitude' in header_map else None,
                    'longitude': parse_float(row.get(header_map['longitude'])) if 'longitude' in header_map else None
                })
        
        # Clear only this route's database records to support multiple routes
        db.query(Cliente).filter(Cliente.rota == rota).delete()
        
        # Save to DB
        for item in rows_to_process:
            cod_val = item['cod_cliente']
            if not cod_val:
                continue
                
            cliente = Cliente(
                cod_cliente=cod_val,
                razao_social=item['razao_social'],
                endereco=item['endereco'],
                bairro=item['bairro'],
                cidade=item['cidade'],
                canal_resumido=item['canal_resumido'],
                classificacao=item['classificacao'],
                novo_dia=item['novo_dia'],
                nova_semana=item['nova_semana'],
                latitude=item.get('latitude'),
                longitude=item.get('longitude'),
                rota=rota
            )
            db.add(cliente)
        db.commit()
        return {"message": f"Rota {rota} importada com sucesso ({len(rows_to_process)} clientes)!"}
    except Exception as e:
        return JSONResponse(status_code=400, content={"message": f"Erro ao processar arquivo: {str(e)}"})

@app.post("/api/metas")
async def save_metas(
    meta_sempre_juntos_pct: float = Form(...),
    meta_cerveja_total: int = Form(...),
    meta_cerveja_600ml: int = Form(...),
    meta_cerveja_ln: int = Form(...),
    meta_cerveja_lata: int = Form(...),
    meta_artd: int = Form(...),
    meta_monster: int = Form(...),
    meta_perfetti: int = Form(...),
    meta_campari: int = Form(...),
    rota: str = Form(None),
    db: Session = Depends(get_db)
):
    current_month = get_current_month()
    if not rota or rota == "undefined" or rota == "null":
        first_route = db.query(Cliente.rota).filter(Cliente.rota.isnot(None), Cliente.rota != "").order_by(Cliente.rota.asc()).first()
        rota = first_route[0] if first_route else None

    meta = db.query(MetaMensal).filter(
        MetaMensal.mes_ano == current_month,
        MetaMensal.rota == rota
    ).first()
    
    if not meta:
        meta = MetaMensal(mes_ano=current_month, rota=rota)
        db.add(meta)
        
    meta.meta_sempre_juntos_pct = meta_sempre_juntos_pct
    meta.meta_cerveja_total = meta_cerveja_total
    meta.meta_cerveja_600ml = meta_cerveja_600ml
    meta.meta_cerveja_ln = meta_cerveja_ln
    meta.meta_cerveja_lata = meta_cerveja_lata
    meta.meta_artd = meta_artd
    meta.meta_monster = meta_monster
    meta.meta_perfetti = meta_perfetti
    meta.meta_campari = meta_campari
    
    db.commit()
    return {"message": "Metas salvas com sucesso!"}

@app.get("/api/metas")
async def get_metas(rota: str = None, db: Session = Depends(get_db)):
    current_month = get_current_month()
    if not rota or rota == "undefined" or rota == "null":
        first_route = db.query(Cliente.rota).filter(Cliente.rota.isnot(None), Cliente.rota != "").order_by(Cliente.rota.asc()).first()
        rota = first_route[0] if first_route else None

    meta = db.query(MetaMensal).filter(
        MetaMensal.mes_ano == current_month,
        MetaMensal.rota == rota
    ).first()
    if not meta:
        return {}
    return {
        "meta_sempre_juntos_pct": meta.meta_sempre_juntos_pct,
        "meta_cerveja_total": meta.meta_cerveja_total,
        "meta_cerveja_600ml": meta.meta_cerveja_600ml,
        "meta_cerveja_ln": meta.meta_cerveja_ln,
        "meta_cerveja_lata": meta.meta_cerveja_lata,
        "meta_artd": meta.meta_artd,
        "meta_monster": meta.meta_monster,
        "meta_perfetti": meta.meta_perfetti,
        "meta_campari": meta.meta_campari
    }

@app.get("/api/rotas")
async def get_rotas(db: Session = Depends(get_db)):
    rotas_query = db.query(Cliente.rota).distinct().all()
    return sorted([r[0] for r in rotas_query if r[0]])

@app.get("/api/clientes")
async def get_clientes(
    dia: str = None, 
    semana: str = None, 
    status_meta: str = None, 
    user_lat: float = None, 
    user_lng: float = None, 
    rota: str = None,
    background_tasks: BackgroundTasks = None,
    db: Session = Depends(get_db)
):
    if not rota or rota.strip() == "" or rota == "undefined" or rota == "null":
        first_route = db.query(Cliente.rota).filter(Cliente.rota.isnot(None), Cliente.rota != "").order_by(Cliente.rota.asc()).first()
        if first_route:
            rota = first_route[0]
        else:
            rota = None

    query = db.query(Cliente)
    if rota:
        query = query.filter(Cliente.rota == rota)
    if dia:
        query = query.filter(Cliente.novo_dia == dia)
    if semana:
        if semana == 'S1':
            query = query.filter((Cliente.nova_semana == 'S1') | (Cliente.nova_semana.contains('S1S3')))
        elif semana == 'S2':
            query = query.filter((Cliente.nova_semana == 'S2') | (Cliente.nova_semana.contains('S2S4')))
        elif semana == 'S3':
            query = query.filter((Cliente.nova_semana == 'S3') | (Cliente.nova_semana.contains('S1S3')))
        elif semana == 'S4':
            query = query.filter((Cliente.nova_semana == 'S4') | (Cliente.nova_semana.contains('S2S4')))
        else:
            query = query.filter(Cliente.nova_semana == semana)
            
    if status_meta and status_meta != 'todos':
        current_month = get_current_month()
        produto_id = None
        
        if status_meta == 'cervejas':
            prod = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == "Cervejas").first()
            if prod: produto_id = prod.id
        elif status_meta == 'drinks':
            prod = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == "Drinks").first()
            if prod: produto_id = prod.id
        elif status_meta == 'sempre_juntos':
            prod = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == "Sempre Juntos").first()
            if prod: produto_id = prod.id
        else:
            # Check if dynamic launch ID
            try:
                produto_id = int(status_meta)
            except ValueError:
                prod = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == status_meta).first()
                if prod: produto_id = prod.id
                
        if produto_id:
            subquery = db.query(PositivacaoDinamica.cod_cliente).filter(
                PositivacaoDinamica.mes_ano == current_month,
                PositivacaoDinamica.produto_id == produto_id,
                PositivacaoDinamica.valor == True
            ).subquery()
            query = query.filter(Cliente.cod_cliente.in_(subquery))
            
    query = query.order_by(Cliente.razao_social.asc())
    clientes = query.all()
    
    # Progressive Geocoding (queued in background to keep responses instant)
    missing_geo_ids = [c.cod_cliente for c in clientes if c.latitude is None or c.longitude is None]
    if missing_geo_ids and background_tasks:
        from database import SessionLocal
        background_tasks.add_task(geocode_missing_clients, missing_geo_ids[:3], SessionLocal)
    
    current_month = get_current_month()
    pos_checks_query = db.query(PositivacaoDinamica).filter(
        PositivacaoDinamica.mes_ano == current_month,
        PositivacaoDinamica.valor == True
    )
    if rota:
        pos_checks_query = pos_checks_query.filter(PositivacaoDinamica.rota == rota)
    pos_checks = pos_checks_query.all()
    
    products = db.query(ProdutoMeta).all()
    prod_map = {p.id: p.nome_produto for p in products}
    
    client_checks = {}
    for rec in pos_checks:
        prod_name = prod_map.get(rec.produto_id)
        if not prod_name:
            continue
        if rec.cod_cliente not in client_checks:
            client_checks[rec.cod_cliente] = set()
        client_checks[rec.cod_cliente].add(prod_name)
        
    result = []
    for c in clientes:
        checked_names = list(client_checks.get(c.cod_cliente, set()))
        result.append({
            "cod_cliente": c.cod_cliente,
            "razao_social": c.razao_social,
            "endereco": c.endereco,
            "bairro": c.bairro,
            "cidade": c.cidade,
            "canal_resumido": c.canal_resumido,
            "classificacao": c.classificacao,
            "novo_dia": c.novo_dia,
            "nova_semana": c.nova_semana,
            "latitude": c.latitude,
            "longitude": c.longitude,
            "positivados": checked_names
        })
        
    # Geographic Nearest Neighbor reordering
    if user_lat is not None and user_lng is not None:
        have_coords = []
        no_coords = []
        for c in result:
            if c.get("latitude") is not None and c.get("longitude") is not None:
                have_coords.append(c)
            else:
                no_coords.append(c)
                
        ordered_list = []
        current_lat = user_lat
        current_lng = user_lng
        
        # If user coordinates are fallback at (0.0, 0.0) but we have coordinates, use the first client's coords as start point!
        if abs(user_lat) < 0.01 and abs(user_lng) < 0.01 and have_coords:
            current_lat = have_coords[0]["latitude"]
            current_lng = have_coords[0]["longitude"]
            
        while have_coords:
            best_idx = 0
            best_dist = haversine_distance(current_lat, current_lng, have_coords[0]["latitude"], have_coords[0]["longitude"])
            for idx in range(1, len(have_coords)):
                d = haversine_distance(current_lat, current_lng, have_coords[idx]["latitude"], have_coords[idx]["longitude"])
                if d < best_dist:
                    best_dist = d
                    best_idx = idx
            best_client = have_coords.pop(best_idx)
            ordered_list.append(best_client)
            current_lat = best_client["latitude"]
            current_lng = best_client["longitude"]
            
        result = ordered_list + no_coords
        
    return result

@app.get("/api/cliente/{cod_cliente}")
async def get_cliente_data(cod_cliente: str, db: Session = Depends(get_db)):
    cliente = db.query(Cliente).filter(Cliente.cod_cliente == cod_cliente).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
        
    current_month = get_current_month()
    
    # Fetch all active products
    products = db.query(ProdutoMeta).all()
    
    # Fetch positive checks for this client in the current month
    pos_records_query = db.query(PositivacaoDinamica).filter(
        PositivacaoDinamica.cod_cliente == cod_cliente,
        PositivacaoDinamica.mes_ano == current_month,
        PositivacaoDinamica.valor == True
    )
    if cliente and cliente.rota:
        pos_records_query = pos_records_query.filter(PositivacaoDinamica.rota == cliente.rota)
    pos_records = pos_records_query.all()
    
    checked_state = {}
    prod_map = {p.id: p.nome_produto for p in products}
    core_names = ("Cervejas", "Drinks", "Sempre Juntos", "Monster", "Perfetti", "Alcoólicos", "Campari")
    
    for prod in products:
        prod_name = prod.nome_produto
        if prod_name in core_names:
            prod_key = prod_name.lower().replace("á", "a").replace("ó", "o").replace(" ", "_")
            if prod_key == "campari":
                prod_key = "alcoolicos"
            checked_state[prod_key] = False
        else:
            checked_state[f"prod_{prod.id}"] = False
            
    for rec in pos_records:
        prod_name = prod_map.get(rec.produto_id)
        if not prod_name:
            continue
        
        if prod_name in core_names:
            prod_key = prod_name.lower().replace("á", "a").replace("ó", "o").replace(" ", "_")
            if prod_key == "campari":
                prod_key = "alcoolicos"
            
            if rec.sub_item:
                checked_state[f"{prod_key}:{rec.sub_item}"] = True
                checked_state[prod_key] = True
            else:
                checked_state[prod_key] = True
        else:
            checked_state[f"prod_{rec.produto_id}"] = True
            
    return {
        "cliente": {
            "cod_cliente": cliente.cod_cliente,
            "razao_social": cliente.razao_social,
            "endereco": cliente.endereco,
            "bairro": cliente.bairro,
            "cidade": cliente.cidade,
            "canal_resumido": cliente.canal_resumido,
            "classificacao": cliente.classificacao,
            "novo_dia": cliente.novo_dia,
            "nova_semana": cliente.nova_semana
        },
        "positivacao": checked_state
    }

@app.post("/api/positivacao/{cod_cliente}")
async def update_positivacao(cod_cliente: str, request: Request, db: Session = Depends(get_db)):
    content_type = request.headers.get("content-type", "")
    if "application/json" in content_type:
        data = await request.json()
    else:
        form_data = await request.form()
        data = {}
        for k, v in form_data.items():
            data[k] = v in ('on', 'true', '1')
    current_month = get_current_month()
    
    cliente = db.query(Cliente).filter(Cliente.cod_cliente == cod_cliente).first()
    rota = cliente.rota if cliente else None
    
    for key, val in data.items():
        produto_id = None
        sub_item = None
        
        if key == 'sempre_juntos':
            prod = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == "Sempre Juntos").first()
            if prod: produto_id = prod.id
        elif key in ('cervejas', 'cerveja_total'):
            prod = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == "Cervejas").first()
            if prod: produto_id = prod.id
        elif key in ('alcoolicos', 'alcoolicos_total'):
            prod = db.query(ProdutoMeta).filter((ProdutoMeta.nome_produto == "Alcoólicos") | (ProdutoMeta.nome_produto == "Campari")).first()
            if prod: produto_id = prod.id
        elif key in ('drinks', 'drinks_total'):
            prod = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == "Drinks").first()
            if prod: produto_id = prod.id
        elif key == 'monster':
            prod = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == "Monster").first()
            if prod: produto_id = prod.id
        elif key == 'perfetti':
            prod = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == "Perfetti").first()
            if prod: produto_id = prod.id
        elif ":" in key:
            prefix, sub_name = key.split(":", 1)
            prod_name = None
            if prefix == 'cervejas':
                prod_name = "Cervejas"
            elif prefix == 'drinks':
                prod_name = "Drinks"
            elif prefix == 'alcoolicos':
                prod_name = "Alcoólicos"
                
            if prod_name:
                prod = db.query(ProdutoMeta).filter((ProdutoMeta.nome_produto == prod_name) | (ProdutoMeta.nome_produto == "Campari" if prod_name == "Alcoólicos" else False)).first()
                if prod:
                    produto_id = prod.id
                    sub_item = sub_name.strip()
        elif key.startswith('prod_'):
            try:
                produto_id = int(key.replace('prod_', ''))
            except ValueError:
                continue
                
        if produto_id is not None:
            record = db.query(PositivacaoDinamica).filter(
                PositivacaoDinamica.cod_cliente == cod_cliente,
                PositivacaoDinamica.mes_ano == current_month,
                PositivacaoDinamica.produto_id == produto_id,
                PositivacaoDinamica.sub_item == sub_item
            ).first()
            
            if not record:
                record = PositivacaoDinamica(
                    cod_cliente=cod_cliente,
                    mes_ano=current_month,
                    produto_id=produto_id,
                    sub_item=sub_item,
                    mes_referencia=datetime.now().strftime("%m/%Y"),
                    data_registro=datetime.now(),
                    rota=rota
                )
                db.add(record)
            else:
                record.mes_referencia = datetime.now().strftime("%m/%Y")
                record.data_registro = datetime.now()
                record.rota = rota
                
            record.valor = bool(val)
            
            # If master was unchecked, uncheck all sub-items under this product
            if key in ('cervejas', 'cerveja_total', 'alcoolicos', 'alcoolicos_total', 'drinks', 'drinks_total') and not val:
                db.query(PositivacaoDinamica).filter(
                    PositivacaoDinamica.cod_cliente == cod_cliente,
                    PositivacaoDinamica.mes_ano == current_month,
                    PositivacaoDinamica.produto_id == produto_id
                ).update({PositivacaoDinamica.valor: False})
                
    db.commit()
    return {"message": "Salvo"}

@app.get("/api/produtos")
async def get_produtos(db: Session = Depends(get_db)):
    return db.query(ProdutoMeta).all()

@app.post("/api/produtos")
async def add_produto(nome_produto: str = Form(...), meta_quantidade: int = Form(...), db: Session = Depends(get_db)):
    nome_clean = nome_produto.strip()
    if not nome_clean:
        return JSONResponse(status_code=400, content={"message": "Nome do produto não pode ser vazio."})
        
    exists = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == nome_clean).first()
    if exists:
        return JSONResponse(status_code=400, content={"message": "Este produto já existe."})
        
    prod = ProdutoMeta(nome_produto=nome_clean, meta_quantidade=meta_quantidade, obrigatorio_sempre_juntos=False)
    db.add(prod)
    db.commit()
    return {
        "message": f"Produto '{nome_clean}' adicionado com sucesso!",
        "produto": {"id": prod.id, "nome_produto": prod.nome_produto, "meta": prod.meta_quantidade}
    }

@app.delete("/gerenciamento/deletar/{produto_id}")
@app.post("/gerenciamento/deletar/{produto_id}")
async def delete_produto(produto_id: int, db: Session = Depends(get_db)):
    prod = db.query(ProdutoMeta).filter(ProdutoMeta.id == produto_id).first()
    if not prod:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
        
    # Also delete associated positivacoes
    db.query(PositivacaoDinamica).filter(PositivacaoDinamica.produto_id == produto_id).delete()
    
    db.delete(prod)
    db.commit()
    return HTMLResponse(content="", status_code=200)

@app.get("/api/dashboard")
async def get_dashboard(rota: str = None, db: Session = Depends(get_db)):
    current_month = get_current_month()
    
    if not rota or rota.strip() == "" or rota == "undefined" or rota == "null":
        first_route = db.query(Cliente.rota).filter(Cliente.rota.isnot(None), Cliente.rota != "").order_by(Cliente.rota.asc()).first()
        rota = first_route[0] if first_route else None
        
    meta = db.query(MetaMensal).filter(
        MetaMensal.mes_ano == current_month,
        MetaMensal.rota == rota
    ).first()
    
    if not meta:
        # Create default baseline meta to prevent crashes
        meta = MetaMensal(
            mes_ano=current_month,
            rota=rota,
            meta_sempre_juntos_pct=39.0,
            meta_cerveja_total=10,
            meta_cerveja_600ml=10,
            meta_cerveja_ln=10,
            meta_cerveja_lata=10,
            meta_artd=10,
            meta_monster=10,
            meta_perfetti=10,
            meta_campari=10
        )
        db.add(meta)
        db.commit()
        
    query_clients = db.query(Cliente)
    if rota:
        query_clients = query_clients.filter(Cliente.rota == rota)
        
    total_clients = query_clients.count()
    canais_validos = ['BAR', 'LANCHONETES', 'RESTAURANTE', 'PADARIA', 'MERCEARIA', 'Bar', 'Lanchonete', 'Restaurante', 'Padaria', 'Mercearia']
    clientes_validos_count = query_clients.filter(Cliente.canal_resumido.in_(canais_validos)).count()
    
    products = db.query(ProdutoMeta).all()
    
    realizado = {}
    metas_dict = {}
    
    prod_sj = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == "Sempre Juntos").first()
    prod_cervejas = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == "Cervejas").first()
    prod_drinks = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == "Drinks").first()
    
    if prod_sj:
        sj_query = db.query(PositivacaoDinamica).join(
            Cliente, Cliente.cod_cliente == PositivacaoDinamica.cod_cliente
        ).filter(
            PositivacaoDinamica.mes_ano == current_month,
            PositivacaoDinamica.produto_id == prod_sj.id,
            PositivacaoDinamica.valor == True,
            Cliente.canal_resumido.in_(canais_validos)
        )
        if rota:
            sj_query = sj_query.filter(Cliente.rota == rota)
            
        sj_count = sj_query.count()
        sj_pct = round((sj_count / clientes_validos_count * 100), 1) if clientes_validos_count > 0 else 0.0
        realizado["sempre_juntos_pct"] = sj_pct
        metas_dict["sempre_juntos_pct"] = meta.meta_sempre_juntos_pct
        
    if prod_cervejas:
        cv_query = db.query(PositivacaoDinamica.cod_cliente).join(
            Cliente, Cliente.cod_cliente == PositivacaoDinamica.cod_cliente
        ).filter(
            PositivacaoDinamica.mes_ano == current_month,
            PositivacaoDinamica.produto_id == prod_cervejas.id,
            PositivacaoDinamica.valor == True
        )
        if rota:
            cv_query = cv_query.filter(Cliente.rota == rota)
            
        cv_count = cv_query.distinct().count()
        realizado["cerveja_total"] = cv_count
        metas_dict["cerveja_total"] = meta.meta_cerveja_total
        
        # Pull sub-items positive checks to aggregate dynamically
        sub_query = db.query(PositivacaoDinamica).join(
            Cliente, Cliente.cod_cliente == PositivacaoDinamica.cod_cliente
        ).filter(
            PositivacaoDinamica.mes_ano == current_month,
            PositivacaoDinamica.produto_id == prod_cervejas.id,
            PositivacaoDinamica.valor == True
        )
        if rota:
            sub_query = sub_query.filter(Cliente.rota == rota)
            
        cervejas_pos = sub_query.all()
        c_600ml_clients = set()
        c_ln_clients = set()
        c_lata_clients = set()
        
        for p_rec in cervejas_pos:
            sub = (p_rec.sub_item or '').lower()
            if not sub:
                continue
            if '600ml' in sub or '500ml' in sub or 'vidro' in sub:
                c_600ml_clients.add(p_rec.cod_cliente)
            elif 'ln' in sub or 'long neck' in sub:
                c_ln_clients.add(p_rec.cod_cliente)
            elif 'lata' in sub:
                c_lata_clients.add(p_rec.cod_cliente)
                
        realizado["cerveja_600ml"] = len(c_600ml_clients)
        realizado["cerveja_ln"] = len(c_ln_clients)
        realizado["cerveja_lata"] = len(c_lata_clients)
        
        metas_dict["cerveja_600ml"] = meta.meta_cerveja_600ml
        metas_dict["cerveja_ln"] = meta.meta_cerveja_ln
        metas_dict["cerveja_lata"] = meta.meta_cerveja_lata
            
    if prod_drinks:
        drinks_query = db.query(PositivacaoDinamica.cod_cliente).join(
            Cliente, Cliente.cod_cliente == PositivacaoDinamica.cod_cliente
        ).filter(
            PositivacaoDinamica.mes_ano == current_month,
            PositivacaoDinamica.produto_id == prod_drinks.id,
            PositivacaoDinamica.valor == True
        )
        if rota:
            drinks_query = drinks_query.filter(Cliente.rota == rota)
            
        drinks_count = drinks_query.distinct().count()
        realizado["drinks"] = drinks_count
        metas_dict["drinks"] = meta.meta_artd
        
    prod_monster = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == "Monster").first()
    if prod_monster:
        monster_query = db.query(PositivacaoDinamica.cod_cliente).join(
            Cliente, Cliente.cod_cliente == PositivacaoDinamica.cod_cliente
        ).filter(
            PositivacaoDinamica.mes_ano == current_month,
            PositivacaoDinamica.produto_id == prod_monster.id,
            PositivacaoDinamica.valor == True
        )
        if rota:
            monster_query = monster_query.filter(Cliente.rota == rota)
            
        monster_count = monster_query.distinct().count()
        realizado["monster"] = monster_count
        metas_dict["monster"] = meta.meta_monster
        
    prod_perfetti = db.query(ProdutoMeta).filter(ProdutoMeta.nome_produto == "Perfetti").first()
    if prod_perfetti:
        perfetti_query = db.query(PositivacaoDinamica.cod_cliente).join(
            Cliente, Cliente.cod_cliente == PositivacaoDinamica.cod_cliente
        ).filter(
            PositivacaoDinamica.mes_ano == current_month,
            PositivacaoDinamica.produto_id == prod_perfetti.id,
            PositivacaoDinamica.valor == True
        )
        if rota:
            perfetti_query = perfetti_query.filter(Cliente.rota == rota)
            
        perfetti_count = perfetti_query.distinct().count()
        realizado["perfetti"] = perfetti_count
        metas_dict["perfetti"] = meta.meta_perfetti
        
    prod_alcoolicos = db.query(ProdutoMeta).filter((ProdutoMeta.nome_produto == "Alcoólicos") | (ProdutoMeta.nome_produto == "Campari")).first()
    if prod_alcoolicos:
        alcoolicos_query = db.query(PositivacaoDinamica.cod_cliente).join(
            Cliente, Cliente.cod_cliente == PositivacaoDinamica.cod_cliente
        ).filter(
            PositivacaoDinamica.mes_ano == current_month,
            PositivacaoDinamica.produto_id == prod_alcoolicos.id,
            PositivacaoDinamica.valor == True
        )
        if rota:
            alcoolicos_query = alcoolicos_query.filter(Cliente.rota == rota)
            
        alcoolicos_count = alcoolicos_query.distinct().count()
        realizado["campari"] = alcoolicos_count
        metas_dict["campari"] = meta.meta_campari
        
    dynamic_products = [p for p in products if p.nome_produto not in ("Cervejas", "Drinks", "Sempre Juntos", "Monster", "Perfetti", "Campari", "Alcoólicos")]
    launches_realizado = []
    
    for lp in dynamic_products:
        lp_query = db.query(PositivacaoDinamica.cod_cliente).join(
            Cliente, Cliente.cod_cliente == PositivacaoDinamica.cod_cliente
        ).filter(
            PositivacaoDinamica.mes_ano == current_month,
            PositivacaoDinamica.produto_id == lp.id,
            PositivacaoDinamica.valor == True
        )
        if rota:
            lp_query = lp_query.filter(Cliente.rota == rota)
            
        count = lp_query.distinct().count()
        
        launches_realizado.append({
            "id": lp.id,
            "nome_produto": lp.nome_produto,
            "realizado": count,
            "meta": lp.meta_quantidade or 10
        })
        
    # SKU positive-checked client count query
    from sqlalchemy import func
    skus_realizado = {}
    sku_query = db.query(
        PositivacaoDinamica.produto_id,
        PositivacaoDinamica.sub_item,
        func.count(PositivacaoDinamica.cod_cliente.distinct())
    ).join(
        Cliente, Cliente.cod_cliente == PositivacaoDinamica.cod_cliente
    ).filter(
        PositivacaoDinamica.mes_ano == current_month,
        PositivacaoDinamica.valor == True,
        PositivacaoDinamica.sub_item.isnot(None),
        PositivacaoDinamica.sub_item != ""
    )
    if rota:
        sku_query = sku_query.filter(Cliente.rota == rota)
        
    sku_results = sku_query.group_by(PositivacaoDinamica.produto_id, PositivacaoDinamica.sub_item).all()
    
    # Map products
    for pid, sub_name, count in sku_results:
        prod_obj = db.query(ProdutoMeta).filter(ProdutoMeta.id == pid).first()
        if prod_obj:
            pname = prod_obj.nome_produto
            prefix = None
            if pname == "Cervejas":
                prefix = "cervejas"
            elif pname in ("Alcoólicos", "Campari"):
                prefix = "alcoolicos"
            elif pname == "Drinks":
                prefix = "drinks"
            if prefix:
                skus_realizado[f"{prefix}:{sub_name}"] = count

    return {
        "metas": metas_dict,
        "realizado": realizado,
        "launches": launches_realizado,
        "skus": skus_realizado,
        "total_clientes": total_clients,
        "clientes_validos_sj": clientes_validos_count
    }

@app.post("/api/chat")
async def post_chat(
    mensagem: str = Form(...),
    active_route: str = Form(None),
    db: Session = Depends(get_db)
):
    if not active_route or active_route == "undefined" or active_route == "null":
        first_route = db.query(Cliente.rota).filter(Cliente.rota.isnot(None), Cliente.rota != "").order_by(Cliente.rota.asc()).first()
        active_route = first_route[0] if first_route else "Geral"
        
    # Save user message to database
    user_msg = HistoricoChat(autor="user", texto=mensagem, rota_ativa=active_route)
    db.add(user_msg)
    db.commit()
    
    # 1. Busca do Mês Atual e Anterior
    curr_yyyymm, prev_mmyyyy = get_current_and_prev_months()
    
    clientes = db.query(Cliente).filter(Cliente.rota == active_route).all()
    cod_clientes = [c.cod_cliente for c in clientes]
    
    # 2. Busca do Mês Atual
    curr_pos = db.query(PositivacaoDinamica).filter(
        PositivacaoDinamica.cod_cliente.in_(cod_clientes),
        PositivacaoDinamica.mes_ano == curr_yyyymm
    ).all()
    
    products = db.query(ProdutoMeta).all()
    prod_map = {p.id: p.nome_produto for p in products}
    
    curr_state = {}
    for p in curr_pos:
        pname = prod_map.get(p.produto_id)
        if pname:
            curr_state[(p.cod_cliente, pname, p.sub_item)] = p.valor
            
    # Meta
    meta = db.query(MetaMensal).filter(
        MetaMensal.mes_ano == curr_yyyymm,
        MetaMensal.rota == active_route
    ).first()
    
    metas_str = "Nenhuma meta cadastrada"
    if meta:
        metas_str = (
            f"Sempre Juntos: {meta.meta_sempre_juntos_pct}%, "
            f"Cervejas Total: {meta.meta_cerveja_total}, "
            f"Cerveja 600ml: {meta.meta_cerveja_600ml}, "
            f"Long Neck: {meta.meta_cerveja_ln}, "
            f"Lata: {meta.meta_cerveja_lata}, "
            f"Drinks: {meta.meta_artd}, "
            f"Monster: {meta.meta_monster}, "
            f"Perfetti: {meta.meta_perfetti}, "
            f"Alcoólicos: {meta.meta_campari}"
        )
        
    # 3. Busca do Mês Anterior
    prev_history = db.query(HistoricoPositivacaoMensal).filter(
        HistoricoPositivacaoMensal.cod_cliente.in_(cod_clientes),
        HistoricoPositivacaoMensal.mes_ano == prev_mmyyyy
    ).all()
    
    # Cruzamento (Churn/Queda de Mix)
    alerts = []
    for h in prev_history:
        if h.positivado:
            current_val = curr_state.get((h.cod_cliente, h.categoria_principal, h.sku_especifico))
            if not current_val:
                cliente = next((c for c in clientes if c.cod_cliente == h.cod_cliente), None)
                razao_social = cliente.razao_social if cliente else f"Cliente #{h.cod_cliente}"
                prod_label = h.sku_especifico if h.sku_especifico else h.categoria_principal
                
                alerts.append(
                    f"- {razao_social} (Cod: {h.cod_cliente}): comprou {prod_label} no mês passado ({prev_mmyyyy}), mas NÃO comprou no mês atual ({curr_yyyymm})."
                )
    alerts_str = "\n".join(alerts) if alerts else "Nenhum alerta de queda de mix."
    
    # Client summary with channel and classification
    client_list = []
    for c in clientes:
        bought_this_month = []
        for k, v in curr_state.items():
            if k[0] == c.cod_cliente and v == True:
                prod_name = f"{k[1]} ({k[2]})" if k[2] else k[1]
                bought_this_month.append(prod_name)
        bought_str = ", ".join(bought_this_month) if bought_this_month else "Nenhum produto"
        
        client_list.append(
            f"Cod: {c.cod_cliente} | {c.razao_social} | Canal: {c.canal_resumido} | Classificação: {c.classificacao} | Visita: {c.novo_dia} ({c.nova_semana}) | Positivado este mês: [{bought_str}]"
        )
    clients_str = "\n".join(client_list)
    
    # 4. Chat history
    chat_history_records = db.query(HistoricoChat).filter(
        HistoricoChat.rota_ativa == active_route
    ).order_by(HistoricoChat.data_hora.desc()).limit(10).all()
    chat_history_records.reverse()
    
    # Call OpenAI
    system_prompt = f"""Você é um Gerente de Vendas Sênior da Coca-Cola Andina, focado na região de Campo Grande, Rio de Janeiro. 
Seu papel absoluto é guiar o vendedor Carlos para bater suas metas mensais e colocar o máximo de comissão no bolso através de análises de campo certeiras e estratégicas.

Tom de Voz:
- Direto, extremamente focado em vendas e metas numéricas.
- Linguagem comercial, motivadora, mas sem enrolação ou rodeios. Escreva de forma objetiva, em parágrafos curtos ou tópicos de fácil leitura em dispositivos móveis.
- Chame-o de Carlos.

Dados de Campo Disponíveis da Rota Ativa ({active_route}):
1. Metas da Rota para este mês ({curr_yyyymm}):
{metas_str}

2. Lista de Clientes da Rota (contendo canal, classificação, dia de visita e positivações atuais do mês):
{clients_str}

3. Alertas de Churn / Queda de Mix (clientes que compraram no mês passado {prev_mmyyyy} mas ainda estão zerados neste mês {curr_yyyymm}):
{alerts_str}

Capacidades Analíticas:
1. Oportunidades por Canal: Se o Carlos perguntar sobre oportunidades de Sempre Juntos ou outras metas por canais ou dias de visita (ex: botecos na S3), consulte a lista de clientes acima, filtre os clientes elegíveis (canais BAR, LANCHONETES, RESTAURANTE, PADARIA, MERCEARIA, etc.) que ainda não positivaram a respectiva categoria ou SKU este mês, e liste os principais alvos estratégicos com nome e código.
2. Comparação de Perdas: Se o Carlos perguntar quem comprou mês passado e não comprou este mês, responda com base nos Alertas de Churn / Queda de Mix listados acima, detalhando o nome do cliente, o código e qual produto/embalagem ele deixou de comprar.
"""

    messages = [{"role": "system", "content": system_prompt}]
    for msg in chat_history_records[:-1]: # Exclude the user message we just saved to put it in chronological order at the end
        role = "user" if msg.autor == "user" else "assistant"
        messages.append({"role": role, "content": msg.texto})
    # Add the current message
    messages.append({"role": "user", "content": mensagem})
        
    if not openai_client:
        # Generate smart mock response
        if "comprou" in mensagem.lower() or "perda" in mensagem.lower() or "churn" in mensagem.lower() or "passado" in mensagem.lower():
            if alerts:
                mock_text = f"Fala Carlos! Identifiquei os seguintes clientes em alerta de perda de mix na rota {active_route} (compraram no mês passado e não compraram este mês):\n\n"
                for a in alerts[:4]:
                    mock_text += f"{a}\n"
                mock_text += "\nRecuperar esses volumes é dinheiro rápido no seu bolso. Visite-os com foco hoje! 🚀"
            else:
                mock_text = f"Fala Carlos! Não encontrei nenhum cliente que comprou no mês passado e está zerado neste mês na rota {active_route}. Excelente trabalho mantendo o mix! 🏆"
        elif "oportunidade" in mensagem.lower() or "meta" in mensagem.lower() or "sempre juntos" in mensagem.lower() or "alvo" in mensagem.lower() or "canal" in mensagem.lower():
            # Generate mock target opportunity listing
            botecos = [c for c in clientes if (c.canal_resumido or '').upper() in ('BAR', 'LANCHONETES', 'Bar', 'Lanchonete')]
            unpositivated_sj = []
            for b in botecos:
                has_sj = curr_state.get((b.cod_cliente, "Sempre Juntos", None))
                if not has_sj:
                    unpositivated_sj.append(b)
            if unpositivated_sj:
                mock_text = f"Carlos, temos oportunidades de **Sempre Juntos** no canal Bar/Lanchonetes na rota {active_route}:\n\n"
                for u in unpositivated_sj[:3]:
                    mock_text += f"• **{u.razao_social}** (Cod: {u.cod_cliente}) - Visita: {u.novo_dia} ({u.nova_semana})\n"
                mock_text += f"\nEstes clientes ainda não compraram Sempre Juntos este mês. Vá para a venda agressiva! 💵"
            else:
                mock_text = f"Carlos, fiz a varredura e todos os bares da rota {active_route} já estão positivados em Sempre Juntos este mês. Vamos focar nos demais lançamentos! 🚀"
        else:
            mock_text = (
                f"Fala Carlos! Para habilitar a inteligência artificial preditiva real de Campo Grande, configure a chave `OPENAI_API_KEY` nas variáveis de ambiente do Render.\n\n"
                f"Enquanto isso, fiz a varredura da sua rota ativa **{active_route}** no banco de dados local. Você tem **{len(clientes)}** clientes. "
                f"Use as metas cadastradas e bora bater o mês! 💵"
            )
            
        ia_msg = HistoricoChat(autor="ia", texto=mock_text, rota_ativa=active_route)
        db.add(ia_msg)
        db.commit()
        return HTMLResponse(content=render_chat_bubbles(mensagem, mock_text))
        
    try:
        response = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=messages,
            temperature=0.7
        )
        ia_text = response.choices[0].message.content
        
        ia_msg = HistoricoChat(autor="ia", texto=ia_text, rota_ativa=active_route)
        db.add(ia_msg)
        db.commit()
        
        return HTMLResponse(content=render_chat_bubbles(mensagem, ia_text))
    except Exception as e:
        error_text = f"Erro ao obter resposta da OpenAI: {str(e)}"
        ia_msg = HistoricoChat(autor="ia", texto=error_text, rota_ativa=active_route)
        db.add(ia_msg)
        db.commit()
        return HTMLResponse(content=render_chat_bubbles(mensagem, error_text))

@app.post("/api/chat/limpar")
async def limpar_chat(rota: str = None, db: Session = Depends(get_db)):
    if not rota or rota == "undefined" or rota == "null":
        return JSONResponse(status_code=400, content={"message": "Rota não especificada"})
    db.query(HistoricoChat).filter(HistoricoChat.rota_ativa == rota).delete()
    db.commit()
    return {"message": "Histórico limpo"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=5000, reload=True)
